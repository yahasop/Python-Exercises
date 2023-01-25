import pycountry
from openpyxl import Workbook
import random

# La variable names creara la lista con los nombres en la lista de texto
file1 = open("projectExcel\list.txt")
names = file1.read().splitlines()
namesLength = len(names) # Tome como base el len de names para hacer los ciclos for

# Esto creara la lista con los paises a partir de la libreria pycountry
countryList = []
for i in range(namesLength):
    randCountry = random.choice(list(pycountry.countries))
    countryList.append(randCountry.name)

# Creacion de la lista de los numeros con los premios pero aun no pude hacer que la suma fuera $1,000,000 exactamente
numbersList = []
total = 1000000
numbersList = []
suma = 0
rango = total - suma
while suma < 1000000:
    for i in range(namesLength):
        numbersList.append(random.randint(1,rango))
        suma += numbersList[i]

#Aqui inicia la creacion de la hoja de excel y la insercion de los datos
workbooks = Workbook()
dest_filename = 'projectExcel/excelPython.xlsx'

ws1 = workbooks.active
ws1.title = "excel_project"

ws1.cell(row=1, column=1, value="NAME")
ws1.cell(row=1, column=2, value="PRIZE")
ws1.cell(row=1, column=3, value="COUNTRY")

for i in range(namesLength):
    ws1.cell(row=i+2, column=1, value=names[i])
    ws1.cell(row=i+2, column=2, value=numbersList[i])
    ws1.cell(row=i+2, column=3, value=countryList[i])

workbooks.save(dest_filename)

print("OK") # Simple mensaje para imprimir en la consola
