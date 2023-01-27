from whoisapi import *
from openpyxl import Workbook

import os
import time

# This will read contents from the IP's text file, splitting in lines a converting those lines into strings -
# and then into a list
ips_file = open("projectAPI/ips.txt")
ips_names = ips_file.read().splitlines()

# Main function for the API to be fed with the IP's from the ips_names list. This will create a loop that iterates 
# X numbers of times where X is the IP's in the list and then a try and except if some IP's that are not valid wont
# crash the loop. In this same loop I create the files in a folder called JSON (but the files are plain txt)
for i in range(len(ips_names)):
    time.sleep(1)
    try:
        client = Client(api_key='at_R9HKO0HCoo1Vuu4OpKts3Ds2hLKzV')

        whois = client.data(f'{ips_names[i]}')

        #client.parameters.output_format = 'json'

        json_whois = client.raw_data(f'{ips_names[i]}')

        whois_dn_format = whois.domain_name.replace(".", "-")

        with open(f"projectAPI/JSON/{whois_dn_format}.txt","w") as txt_f:
            #json_f.write(json_whois) This will output a JSON file format
            txt_f.write(f"{str(whois.domain_name)}\n")
            txt_f.write(f"{str(whois.registrar_name)}\n")
            txt_f.write(f"{str(whois.contact_email)}\n")
            txt_f.write(f"{str(whois.registry_data.created_date)}\n")
            txt_f.write(f"{str(whois.registry_data.registrant.country)}\n")

    except KeyError:
        print("Couldn't retrieved the information")

# Here I intialized a variable with the path to the folder where the files are being created
data_files_directory = "projectAPI/JSON"

# I will append the contents of the files in lists and this lists in another list to create a list of lists
ips_data_values = []
# This will check all files within the directory stated in the variable to check their contents and append them to lists
for i in os.listdir(data_files_directory): #This method extracts the file names in the specified directory
    data_file = open(f"{data_files_directory}/{i}", "r")
    data_properties = data_file.read().splitlines()
    # This is what will append the lists into a single list called ips_data_values
    ips_data_values.append(data_properties)
    data_file.close()

# Here is where I append the contents to the Excel workbook
workbook = Workbook() # Workbook class from openpyxl library
wb_dest_filename = 'projectAPI/whois.xlsx'

wb1 = workbook.active
wb1.title = "IPs_data"

# This will print the column names in the whois.xlsx file
wb1.cell(row=1, column=1, value="path filename")
wb1.cell(row=1, column=2, value="domainName")
wb1.cell(row=1, column=3, value="registrarName")
wb1.cell(row=1, column=4, value="contactEmail")
wb1.cell(row=1, column=5, value="registryData.createdDate")
wb1.cell(row=1, column=6, value="registrant.country")

# This loop will generate a list with the names of the files in the ip's data directory
path_file_names = []
for filename in os.listdir(data_files_directory): #This method extracts the file names in the specified directory
    path_file_names.append(filename)

# In this loop, the contents from the ips_data_values and path_file names lists will be written into the specified 
# cells of the workbook. I used nested indexes as the ips_data_values is a list of lists 
for k in range(len(ips_data_values)):
    wb1.cell(row=k+2, column=1, value=path_file_names[k])
    wb1.cell(row=k+2, column=2, value=ips_data_values[k][0])
    wb1.cell(row=k+2, column=3, value=ips_data_values[k][1])
    wb1.cell(row=k+2, column=4, value=ips_data_values[k][2])
    wb1.cell(row=k+2, column=5, value=ips_data_values[k][3])
    wb1.cell(row=k+2, column=6, value=ips_data_values[k][4])

workbook.save(wb_dest_filename)
print("OK")
